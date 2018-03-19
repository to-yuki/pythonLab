# -*- coding: UTF-8 -*-
import openpyxl
import requests
import re
 
# Excel ファイルをロード  
book = openpyxl.load_workbook(filename="GoogleTransExcel.xlsx",data_only=True)
# Excel シート名のリスト取得
sheetnames = book.get_sheet_names()
# 1番目のシートにアクセス
sheet = book.get_sheet_by_name(sheetnames[0])

list = []

for r in range(1,sheet.max_row+1):
    data = sheet.cell(row=r,column=1).value
    list.append(data)

# 正規表現文字列のコンパイル
pattern = re.compile(r"TRANSLATED_TEXT='(.*?)'")
# GoogleTranslateURL
url = u'https://translate.google.com/?hl=ja#en/ja/'
#行カウンタのリセット
r=0

for origin in list:
    if origin == None:
        continue;
    try:
        
        response = requests.get(url, params={'q': origin})
        # resopnse 内にある JavaScript コードから翻訳された文字列を取得
        transValue = pattern.search(response.text).group(1)
        # ExcelSheetに書き込み
        sheet.cell(row=r+1,column=2).value = transValue

        print(u"\n[translate.google.com] で翻訳") 
        print(u"英語:"), 
        print(origin)
        print(u"日本語:"), 
        print(transValue)
        r+=1
    except:
        print(str(r)+u"行目がエラーになりました。"),
# ワークブックの保存
book.save("GoogleTransExcel-OK.xlsx")
